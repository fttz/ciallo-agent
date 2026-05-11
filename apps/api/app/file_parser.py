import csv
import json
from io import StringIO
from pathlib import Path
from urllib.parse import urlparse
import requests
from fastapi import UploadFile
from langchain_community.document_loaders import BSHTMLLoader, PyPDFLoader, TextLoader, UnstructuredPowerPointLoader, WebBaseLoader
from langchain_community.document_loaders.word_document import Docx2txtLoader, UnstructuredWordDocumentLoader
from openpyxl import load_workbook


TEXT_SUFFIXES = {".txt", ".md", ".log"}
HTML_SUFFIXES = {".html", ".htm"}
WORD_SUFFIXES = {".doc", ".docx"}
PPT_SUFFIXES = {".ppt", ".pptx"}
SPREADSHEET_SUFFIXES = {".xlsx"}
WEB_LINK_SUFFIXES = {".url", ".webloc", ".web"}
TABLE_MAX_ROWS = 120
TABLE_MAX_COLUMNS = 20


def _truncate(text: str, limit: int) -> str:
    cleaned = text.strip()
    if len(cleaned) <= limit:
        return cleaned
    return f"{cleaned[:limit]}\n\n[...内容过长，已截断...]"


def _normalize_documents(documents: list, max_chars: int) -> str:
    chunks: list[str] = []
    for index, document in enumerate(documents):
        text = getattr(document, "page_content", "")
        text = "\n".join(line.rstrip() for line in str(text).splitlines() if line.strip())
        if not text:
            continue
        metadata = getattr(document, "metadata", {}) or {}
        page = metadata.get("page")
        source = metadata.get("source")
        label_parts: list[str] = []
        if page is not None:
            label_parts.append(f"page={page}")
        if source and index == 0:
            label_parts.append(f"source={Path(str(source)).name}")
        label = f"[{', '.join(label_parts)}]\n" if label_parts else ""
        chunks.append(f"{label}{text}")

    merged = "\n\n".join(chunks).strip()
    if not merged:
        return "文件解析完成，但未提取到可用文本内容。"
    return _truncate(merged, max_chars)


def _format_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("|", "\\|").strip()
    return text


def _format_table(rows: list[list[object]], max_chars: int) -> str:
    trimmed_rows = [row[:TABLE_MAX_COLUMNS] for row in rows[:TABLE_MAX_ROWS] if any(cell is not None and str(cell).strip() for cell in row)]
    if not trimmed_rows:
        return "文件解析完成，但未提取到可用表格内容。"

    width = max(len(row) for row in trimmed_rows)
    normalized = [row + [""] * (width - len(row)) for row in trimmed_rows]
    header = [_format_cell(cell) or f"列{index + 1}" for index, cell in enumerate(normalized[0])]
    body = normalized[1:] if len(normalized) > 1 else []
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in body:
        lines.append("| " + " | ".join(_format_cell(cell) for cell in row) + " |")
    if len(rows) > TABLE_MAX_ROWS:
        lines.append(f"\n[...表格超过 {TABLE_MAX_ROWS} 行，已截断...]")
    return _truncate("\n".join(lines), max_chars)


async def _parse_csv_upload(file: UploadFile, max_chars: int) -> str:
    raw_text = await _read_upload_text(file)
    sample = raw_text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(StringIO(raw_text), dialect))
    return _format_table(rows, max_chars)


async def _parse_json_upload(file: UploadFile, max_chars: int) -> str:
    raw_text = await _read_upload_text(file)
    parsed = json.loads(raw_text)
    return _truncate(json.dumps(parsed, ensure_ascii=False, indent=2), max_chars)


def _parse_xlsx_file(path: Path, max_chars: int) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sections: list[str] = []
    try:
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            if not rows:
                continue
            sections.append(f"## Sheet: {sheet.title}\n\n{_format_table(rows, max_chars)}")
    finally:
        workbook.close()

    if not sections:
        return "文件解析完成，但未提取到可用表格内容。"
    return _truncate("\n\n".join(sections), max_chars)


def _extract_url_from_text(raw_text: str) -> str:
    for line in raw_text.splitlines():
        candidate = line.strip()
        if not candidate:
            continue
        if candidate.upper().startswith("URL="):
            candidate = candidate[4:].strip()
        parsed = urlparse(candidate)
        if parsed.scheme in {"http", "https"} and parsed.netloc:
            return candidate
    return ""


async def _read_upload_text(file: UploadFile) -> str:
    content = await file.read()
    await file.seek(0)
    for encoding in ("utf-8", "gb18030", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="ignore")


async def save_upload_file(upload_dir: str, file: UploadFile) -> dict:
    base = Path(upload_dir)
    base.mkdir(parents=True, exist_ok=True)

    destination = base / file.filename
    content = await file.read()
    destination.write_bytes(content)
    await file.seek(0)

    return {
        "filename": file.filename,
        "size": len(content),
        "path": str(destination),
    }


async def parse_uploaded_file(upload_dir: str, file: UploadFile, max_chars: int, web_fetch_timeout_sec: int) -> dict:
    saved = await save_upload_file(upload_dir, file)
    path = Path(saved["path"])
    suffix = path.suffix.lower()
    kind = "text"

    try:
        if suffix in TEXT_SUFFIXES:
            loader = TextLoader(str(path), encoding="utf-8", autodetect_encoding=True)
            kind = "text"
        elif suffix == ".csv":
            kind = "sheet"
            content = await _parse_csv_upload(file, max_chars)
            return {
                "filename": saved["filename"],
                "size": saved["size"],
                "kind": kind,
                "content": content,
            }
        elif suffix == ".json":
            kind = "json"
            content = await _parse_json_upload(file, max_chars)
            return {
                "filename": saved["filename"],
                "size": saved["size"],
                "kind": kind,
                "content": content,
            }
        elif suffix in SPREADSHEET_SUFFIXES:
            kind = "sheet"
            content = _parse_xlsx_file(path, max_chars)
            return {
                "filename": saved["filename"],
                "size": saved["size"],
                "kind": kind,
                "content": content,
            }
        elif suffix == ".pdf":
            loader = PyPDFLoader(str(path))
            kind = "pdf"
        elif suffix == ".docx":
            loader = Docx2txtLoader(str(path))
            kind = "word"
        elif suffix == ".doc":
            loader = UnstructuredWordDocumentLoader(str(path))
            kind = "word"
        elif suffix in PPT_SUFFIXES:
            loader = UnstructuredPowerPointLoader(str(path))
            kind = "ppt"
        elif suffix in HTML_SUFFIXES:
            loader = BSHTMLLoader(str(path), open_encoding="utf-8")
            kind = "web"
        elif suffix in WEB_LINK_SUFFIXES:
            kind = "web"
            raw_text = await _read_upload_text(file)
            url = _extract_url_from_text(raw_text)
            if not url:
                raise ValueError("未找到可用的网页链接")
            # Disable inherited proxy settings to avoid local proxy side effects.
            session = requests.Session()
            session.trust_env = False
            loader = WebBaseLoader(
                web_paths=[url],
                requests_per_second=max(1, int(30 / max(web_fetch_timeout_sec, 1))),
                requests_kwargs={"timeout": web_fetch_timeout_sec, "verify": False},
                session=session,
            )
        else:
            raise ValueError("暂不支持该文件类型")

        documents = loader.load()
        content = _normalize_documents(documents, max_chars)
    except Exception as exc:  # noqa: BLE001
        return {
            "filename": saved["filename"],
            "size": saved["size"],
            "kind": kind,
            "content": f"文件解析失败：{exc}",
        }

    return {
        "filename": saved["filename"],
        "size": saved["size"],
        "kind": kind,
        "content": content,
    }
